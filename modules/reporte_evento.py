"""
Reporte post-evento: estadísticas de asistencia, mesas y línea de tiempo.
"""
import customtkinter as ctk
from datetime import datetime
from collections import defaultdict
from modules.database import db

# Usar la versión DPI-aware si ya fue registrada por main.py,
# de lo contrario usar la original (ej. en tests standalone).
_BaseToplevel = ctk.CTkToplevel

COLORS = {
    "bg":         "#0f172a",
    "card":       "#1e293b",
    "border":     "#334155",
    "primary":    "#3b82f6",
    "success":    "#10b981",
    "warning":    "#f59e0b",
    "danger":     "#ef4444",
    "text":       "#f8fafc",
    "text_light": "#94a3b8",
    "gold":       "#d4af37",
}


class ReporteEvento(_BaseToplevel):

    def __init__(self, parent, evento):
        super().__init__(parent)
        self.evento = evento
        self.title(f"Reporte — {evento['nombre']}")
        self.geometry("720x780")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        self.configure(fg_color=COLORS["bg"])

        self.update_idletasks()
        x = (self.winfo_screenwidth()  - 720) // 2
        y = (self.winfo_screenheight() - 780) // 2
        self.geometry(f"+{x}+{y}")

        self._cargar_datos()
        self._build_ui()

    # ------------------------------------------------------------------
    # Datos
    # ------------------------------------------------------------------

    def _cargar_datos(self):
        self.invitados      = db.obtener_invitados_evento(self.evento['id'])
        self.acreditaciones = db.obtener_acreditaciones_evento(self.evento['id'])

        self.total     = len(self.invitados)
        self.presentes = sum(1 for i in self.invitados if i.get('presente'))
        self.ausentes  = self.total - self.presentes
        self.pct       = (self.presentes / self.total * 100) if self.total else 0

        # Agrupar por mesa
        mesas_inv = defaultdict(list)
        for inv in self.invitados:
            mesa = inv.get('mesa') or "Sin mesa"
            mesas_inv[mesa].append(inv)

        self.mesas = []
        for mesa, invs in sorted(mesas_inv.items(), key=lambda x: (str(x[0]).zfill(10))):
            total_m    = len(invs)
            pres_m     = sum(1 for i in invs if i.get('presente'))
            pct_m      = (pres_m / total_m * 100) if total_m else 0
            self.mesas.append({
                "mesa": mesa, "total": total_m,
                "presentes": pres_m, "ausentes": total_m - pres_m,
                "pct": pct_m
            })

        # Línea de tiempo (ingresos agrupados por hora)
        self.timeline = defaultdict(int)
        for acred in self.acreditaciones:
            if acred.get('tipo') == 'ingreso':
                try:
                    ts = datetime.fromisoformat(acred['timestamp'])
                    hora = ts.strftime("%H:%M")
                    # agrupar cada 30 minutos
                    minuto = (ts.minute // 30) * 30
                    clave  = ts.replace(minute=minuto, second=0, microsecond=0).strftime("%H:%M")
                    self.timeline[clave] += 1
                except Exception:
                    pass

        self.primera_acred = None
        self.ultima_acred  = None
        ingresos = [a for a in self.acreditaciones if a.get('tipo') == 'ingreso']
        if ingresos:
            try:
                ordenados = sorted(ingresos, key=lambda a: a['timestamp'])
                self.primera_acred = datetime.fromisoformat(ordenados[0]['timestamp'])
                self.ultima_acred  = datetime.fromisoformat(ordenados[-1]['timestamp'])
            except Exception:
                pass

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color=COLORS["card"],
                              corner_radius=0, height=72)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(header, text=f"📊  Reporte de Evento",
                     font=("Segoe UI", 20, "bold"),
                     text_color=COLORS["gold"]).pack(side="left", padx=20, pady=16)

        ctk.CTkLabel(header, text=self.evento['nombre'],
                     font=("Segoe UI", 14),
                     text_color=COLORS["text_light"]).pack(side="right", padx=20)

        # Scroll principal
        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"])
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        inner = ctk.CTkFrame(scroll, fg_color="transparent")
        inner.pack(fill="x", padx=20, pady=16)

        # ── Resumen general ──────────────────────────────────────────
        self._seccion(inner, "Resumen General")

        resumen = ctk.CTkFrame(inner, fg_color=COLORS["card"], corner_radius=10)
        resumen.pack(fill="x", pady=(0, 16))

        stats_row = ctk.CTkFrame(resumen, fg_color="transparent")
        stats_row.pack(fill="x", padx=20, pady=16)

        self._stat_box(stats_row, str(self.total),     "Total",       COLORS["primary"])
        self._stat_box(stats_row, str(self.presentes), "Asistieron",  COLORS["success"])
        self._stat_box(stats_row, str(self.ausentes),  "Ausentes",    COLORS["danger"])
        self._stat_box(stats_row, f"{self.pct:.0f}%",  "Asistencia",  COLORS["gold"])

        # Barra de asistencia global
        bar_frame = ctk.CTkFrame(resumen, fg_color="transparent")
        bar_frame.pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(bar_frame, text="Asistencia global",
                     font=("Segoe UI", 12),
                     text_color=COLORS["text_light"]).pack(anchor="w")

        bar = ctk.CTkProgressBar(bar_frame, height=16, corner_radius=8,
                                 progress_color=COLORS["success"])
        bar.pack(fill="x", pady=(4, 0))
        bar.set(self.pct / 100)

        # Hora primera / última
        if self.primera_acred and self.ultima_acred:
            info_frame = ctk.CTkFrame(resumen, fg_color="transparent")
            info_frame.pack(fill="x", padx=20, pady=(0, 16))

            duracion = self.ultima_acred - self.primera_acred
            mins     = int(duracion.total_seconds() / 60)
            horas    = mins // 60
            mins_r   = mins % 60
            dur_txt  = f"{horas}h {mins_r}m" if horas else f"{mins_r} min"

            self._info_row(info_frame, "Primera acreditación",
                           self.primera_acred.strftime("%H:%M:%S"))
            self._info_row(info_frame, "Última acreditación",
                           self.ultima_acred.strftime("%H:%M:%S"))
            self._info_row(info_frame, "Duración del evento", dur_txt)

        # ── Línea de tiempo ──────────────────────────────────────────
        if self.timeline:
            self._seccion(inner, "Flujo de Llegadas (cada 30 min)")

            tl_frame = ctk.CTkFrame(inner, fg_color=COLORS["card"], corner_radius=10)
            tl_frame.pack(fill="x", pady=(0, 16))

            max_val = max(self.timeline.values()) or 1

            for hora in sorted(self.timeline.keys()):
                cant = self.timeline[hora]
                row  = ctk.CTkFrame(tl_frame, fg_color="transparent")
                row.pack(fill="x", padx=16, pady=4)

                ctk.CTkLabel(row, text=hora,
                             font=("Segoe UI", 12, "bold"),
                             width=52, anchor="w",
                             text_color=COLORS["text"]).pack(side="left")

                bar_tl = ctk.CTkProgressBar(row, height=18, corner_radius=4,
                                            progress_color=COLORS["primary"])
                bar_tl.pack(side="left", fill="x", expand=True, padx=(8, 8))
                bar_tl.set(cant / max_val)

                ctk.CTkLabel(row, text=str(cant),
                             font=("Segoe UI", 12),
                             width=32, anchor="e",
                             text_color=COLORS["text_light"]).pack(side="left")

        # ── Por mesa ─────────────────────────────────────────────────
        if self.mesas:
            self._seccion(inner, "Por Mesa")

            mesas_frame = ctk.CTkFrame(inner, fg_color=COLORS["card"], corner_radius=10)
            mesas_frame.pack(fill="x", pady=(0, 16))

            # Cabecera
            head = ctk.CTkFrame(mesas_frame, fg_color=COLORS["border"], corner_radius=0)
            head.pack(fill="x")
            for txt, w in [("Mesa", 80), ("Total", 60), ("Presentes", 90),
                           ("Ausentes", 80), ("% Asist.", 80), ("", 0)]:
                ctk.CTkLabel(head, text=txt, font=("Segoe UI", 11, "bold"),
                             text_color=COLORS["text_light"],
                             width=w, anchor="center").pack(side="left", padx=4, pady=6)

            for mesa_data in self.mesas:
                self._fila_mesa(mesas_frame, mesa_data)

        # ── Mesas con 0% ─────────────────────────────────────────────
        ausentes_tot = [m for m in self.mesas if m['presentes'] == 0 and m['mesa'] != "Sin mesa"]
        if ausentes_tot:
            self._seccion(inner, f"Mesas sin ningún asistente ({len(ausentes_tot)})")
            warn = ctk.CTkFrame(inner, fg_color="#2d1b00", corner_radius=10)
            warn.pack(fill="x", pady=(0, 16))

            nombres = ", ".join(f"Mesa {m['mesa']}" for m in ausentes_tot)
            ctk.CTkLabel(warn, text=f"⚠️  {nombres}",
                         font=("Segoe UI", 13),
                         text_color=COLORS["warning"],
                         wraplength=640, justify="left").pack(padx=16, pady=12)

        # ── Botones ──────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(inner, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(8, 0))

        ctk.CTkButton(btn_frame, text="📥 Exportar a Excel",
                      command=self._exportar_excel,
                      height=48, font=("Segoe UI", 14, "bold"),
                      fg_color=COLORS["success"]).pack(side="left", fill="x",
                                                       expand=True, padx=(0, 8))

        ctk.CTkButton(btn_frame, text="Cerrar",
                      command=self.destroy,
                      width=140, height=48, font=("Segoe UI", 14),
                      fg_color="transparent", border_width=2,
                      border_color=COLORS["border"]).pack(side="left")

    # ------------------------------------------------------------------
    # Helpers de UI
    # ------------------------------------------------------------------

    def _seccion(self, parent, titulo):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", pady=(8, 6))
        ctk.CTkLabel(f, text=titulo,
                     font=("Segoe UI", 14, "bold"),
                     text_color=COLORS["text"]).pack(side="left")
        ctk.CTkFrame(f, height=1,
                     fg_color=COLORS["border"]).pack(side="left",
                                                     fill="x", expand=True, padx=(10, 0))

    def _stat_box(self, parent, valor, label, color):
        box = ctk.CTkFrame(parent, fg_color=COLORS["bg"],
                           corner_radius=8)
        box.pack(side="left", fill="x", expand=True, padx=6)
        ctk.CTkLabel(box, text=valor,
                     font=("Segoe UI", 28, "bold"),
                     text_color=color).pack(pady=(12, 2))
        ctk.CTkLabel(box, text=label,
                     font=("Segoe UI", 11),
                     text_color=COLORS["text_light"]).pack(pady=(0, 12))

    def _info_row(self, parent, label, valor):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=2)
        ctk.CTkLabel(row, text=label,
                     font=("Segoe UI", 12),
                     text_color=COLORS["text_light"],
                     width=200, anchor="w").pack(side="left")
        ctk.CTkLabel(row, text=valor,
                     font=("Segoe UI", 12, "bold"),
                     text_color=COLORS["text"]).pack(side="left")

    def _fila_mesa(self, parent, data):
        pct  = data['pct']
        color = (COLORS["success"] if pct >= 80
                 else COLORS["warning"] if pct >= 40
                 else COLORS["danger"])

        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=4, pady=2)

        for txt, w in [(f"Mesa {data['mesa']}", 80),
                       (str(data['total']), 60),
                       (str(data['presentes']), 90),
                       (str(data['ausentes']), 80)]:
            ctk.CTkLabel(row, text=txt,
                         font=("Segoe UI", 12),
                         text_color=COLORS["text"],
                         width=w, anchor="center").pack(side="left", padx=4)

        # Barra % con valor
        bar_wrap = ctk.CTkFrame(row, fg_color="transparent")
        bar_wrap.pack(side="left", fill="x", expand=True, padx=4)

        b = ctk.CTkProgressBar(bar_wrap, height=14, corner_radius=4,
                               progress_color=color)
        b.pack(side="left", fill="x", expand=True)
        b.set(pct / 100)

        ctk.CTkLabel(bar_wrap, text=f"{pct:.0f}%",
                     font=("Segoe UI", 11),
                     text_color=color,
                     width=44, anchor="e").pack(side="left", padx=(6, 0))

    # ------------------------------------------------------------------
    # Exportar Excel
    # ------------------------------------------------------------------

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from tkinter import filedialog
        except ImportError:
            self._msg_error("openpyxl no está instalado.\nEjecutá: pip install openpyxl")
            return

        from tkinter import filedialog
        ruta_trabajo = self.evento.get('ruta_trabajo')
        path = filedialog.asksaveasfilename(
            title="Guardar reporte",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{self.evento['nombre'].replace(' ', '_')}.xlsx",
            initialdir=ruta_trabajo if ruta_trabajo else None
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumen"
        hdr = Font(bold=True, color="FFFFFF")
        fill_azul  = PatternFill("solid", fgColor="1E3A5F")
        fill_verde = PatternFill("solid", fgColor="10b981")
        fill_rojo  = PatternFill("solid", fgColor="ef4444")

        ws1.append(["Reporte de Evento", self.evento['nombre']])
        ws1.append(["Fecha evento", self.evento.get('fecha_evento', '')])
        ws1.append([])
        ws1.append(["Total invitados", self.total])
        ws1.append(["Asistieron",      self.presentes])
        ws1.append(["Ausentes",        self.ausentes])
        ws1.append(["% Asistencia",    f"{self.pct:.1f}%"])
        if self.primera_acred:
            ws1.append(["Primera acreditación", self.primera_acred.strftime("%H:%M:%S")])
            ws1.append(["Última acreditación",  self.ultima_acred.strftime("%H:%M:%S")])

        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 30

        # ── Hoja 2: Por mesa ─────────────────────────────────────────
        ws2 = wb.create_sheet("Por Mesa")
        ws2.append(["Mesa", "Total", "Presentes", "Ausentes", "% Asistencia"])
        for cell in ws2[1]:
            cell.font = hdr
            cell.fill = fill_azul
            cell.alignment = Alignment(horizontal="center")

        for m in self.mesas:
            ws2.append([f"Mesa {m['mesa']}", m['total'],
                        m['presentes'], m['ausentes'], f"{m['pct']:.0f}%"])

        for col in ['A','B','C','D','E']:
            ws2.column_dimensions[col].width = 16

        # ── Hoja 3: Detalle invitados ─────────────────────────────────
        ws3 = wb.create_sheet("Invitados")
        ws3.append(["Nombre", "Apellido", "Mesa", "Presente", "Kiosco"])
        for cell in ws3[1]:
            cell.font = hdr
            cell.fill = fill_azul
            cell.alignment = Alignment(horizontal="center")

        for inv in sorted(self.invitados, key=lambda i: (i.get('mesa') or '', i['apellido'])):
            ws3.append([
                inv['nombre'], inv['apellido'],
                inv.get('mesa') or '',
                "Sí" if inv.get('presente') else "No",
                inv.get('kiosco_acreditador') or ''
            ])

        for col in ['A','B','C','D','E']:
            ws3.column_dimensions[col].width = 20

        # ── Hoja 4: Línea de tiempo ───────────────────────────────────
        if self.timeline:
            ws4 = wb.create_sheet("Flujo")
            ws4.append(["Horario (30 min)", "Acreditaciones"])
            for cell in ws4[1]:
                cell.font = hdr
                cell.fill = fill_azul
                cell.alignment = Alignment(horizontal="center")
            for hora in sorted(self.timeline.keys()):
                ws4.append([hora, self.timeline[hora]])
            ws4.column_dimensions['A'].width = 20
            ws4.column_dimensions['B'].width = 20

        wb.save(path)
        self._msg_ok(f"Reporte exportado:\n{path}")

    def _msg_ok(self, texto):
        d = ctk.CTkToplevel(self)
        d.title("Listo")
        d.geometry("400x160")
        d.transient(self)
        d.grab_set()
        d.update_idletasks()
        x = (d.winfo_screenwidth()  - 400) // 2
        y = (d.winfo_screenheight() - 160) // 2
        d.geometry(f"+{x}+{y}")
        ctk.CTkLabel(d, text="✅  " + texto,
                     font=("Segoe UI", 13),
                     wraplength=360, justify="center").pack(expand=True, pady=20)
        ctk.CTkButton(d, text="OK", command=d.destroy,
                      width=120, height=40).pack(pady=(0, 16))

    def _msg_error(self, texto):
        d = ctk.CTkToplevel(self)
        d.title("Error")
        d.geometry("400x160")
        d.transient(self)
        d.grab_set()
        d.update_idletasks()
        x = (d.winfo_screenwidth()  - 400) // 2
        y = (d.winfo_screenheight() - 160) // 2
        d.geometry(f"+{x}+{y}")
        ctk.CTkLabel(d, text="❌  " + texto,
                     font=("Segoe UI", 13),
                     text_color=COLORS["danger"],
                     wraplength=360, justify="center").pack(expand=True, pady=20)
        ctk.CTkButton(d, text="OK", command=d.destroy,
                      width=120, height=40).pack(pady=(0, 16))
