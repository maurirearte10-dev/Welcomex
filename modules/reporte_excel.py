"""
Generador de Reportes Excel para WelcomeX
Genera reportes completos post-evento
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from datetime import datetime
from pathlib import Path

class GeneradorReporteExcel:
    def __init__(self, evento, invitados, acreditaciones, alertas_seguridad, sorteos):
        self.evento = evento
        self.invitados = invitados
        self.acreditaciones = acreditaciones
        self.alertas = alertas_seguridad
        self.sorteos = sorteos
        self.wb = Workbook()
        
        # Estilos
        self.font_titulo = Font(name='Arial', size=16, bold=True)
        self.font_subtitulo = Font(name='Arial', size=12, bold=True)
        self.font_header = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        self.font_normal = Font(name='Arial', size=10)
        
        self.fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        self.fill_verde = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        self.fill_gris = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        self.fill_rojo = PatternFill(start_color="F5CBA7", end_color="F5CBA7", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generar(self, ruta_salida):
        """Genera el reporte completo"""
        try:
            # Eliminar hoja por defecto
            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])
            
            # Crear hojas
            self.crear_resumen_general()
            self.crear_listado_invitados()
            self.crear_historial_acreditaciones()
            self.crear_alertas_seguridad()
            self.crear_estadisticas_mesas()
            if self.sorteos:
                self.crear_ganadores_sorteos()
            
            # Guardar
            self.wb.save(ruta_salida)
            return {"success": True, "ruta": ruta_salida}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def crear_resumen_general(self):
        """Hoja 1: Resumen General"""
        ws = self.wb.create_sheet("📋 Resumen General", 0)
        
        # Título principal
        ws['A1'] = f"🎉 REPORTE DE EVENTO - {self.evento['nombre'].upper()}"
        ws['A1'].font = self.font_titulo
        ws.merge_cells('A1:D1')
        
        # Info del evento
        row = 3
        ws[f'A{row}'] = f"Fecha: {self.evento.get('fecha_evento', 'N/A')}"
        ws[f'A{row}'].font = self.font_normal
        row += 1
        ws[f'A{row}'] = f"Hora: {self.evento.get('hora_evento', 'N/A')}"
        ws[f'A{row}'].font = self.font_normal
        row += 1
        ws[f'A{row}'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{row}'].font = self.font_normal
        
        # Estadísticas generales
        row += 2
        presentes = len([i for i in self.invitados if i.get('presente')])
        ausentes = len(self.invitados) - presentes
        pct_presentes = (presentes / len(self.invitados) * 100) if self.invitados else 0
        
        ws[f'A{row}'] = "ESTADÍSTICAS GENERALES"
        ws[f'A{row}'].font = self.font_subtitulo
        row += 1
        
        stats = [
            ("Total Invitados:", len(self.invitados)),
            ("Invitados Presentes:", f"{presentes} ({pct_presentes:.1f}%)"),
            ("Invitados Ausentes:", f"{ausentes} ({100-pct_presentes:.1f}%)")
        ]
        
        for label, valor in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = valor
            ws[f'A{row}'].font = self.font_normal
            ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
            row += 1
        
        # Acreditaciones
        row += 1
        ws[f'A{row}'] = "ACREDITACIONES"
        ws[f'A{row}'].font = self.font_subtitulo
        row += 1
        
        ingresos = len([a for a in self.acreditaciones if a.get('tipo') == 'ingreso'])
        egresos = len([a for a in self.acreditaciones if a.get('tipo') == 'egreso'])
        
        ws[f'A{row}'] = "Total Acreditaciones:"
        ws[f'B{row}'] = len(self.acreditaciones)
        ws[f'A{row}'].font = self.font_normal
        ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
        row += 1
        
        ws[f'A{row}'] = "  - Ingresos:"
        ws[f'B{row}'] = ingresos
        row += 1
        
        ws[f'A{row}'] = "  - Egresos:"
        ws[f'B{row}'] = egresos
        row += 1
        
        # Horarios
        if self.acreditaciones:
            row += 1
            ws[f'A{row}'] = "HORARIOS"
            ws[f'A{row}'].font = self.font_subtitulo
            row += 1
            
            # Primera llegada
            primera = min([a for a in self.acreditaciones if a.get('tipo') == 'ingreso'], 
                         key=lambda x: x.get('timestamp', ''), default=None)
            if primera:
                ws[f'A{row}'] = "Primera Llegada:"
                hora = datetime.fromisoformat(primera['timestamp']).strftime('%H:%M')
                ws[f'B{row}'] = f"{hora} ({primera.get('nombre', '')} {primera.get('apellido', '')})"
                row += 1
            
            # Última llegada
            ultima = max([a for a in self.acreditaciones if a.get('tipo') == 'ingreso'], 
                        key=lambda x: x.get('timestamp', ''), default=None)
            if ultima and ultima != primera:
                ws[f'A{row}'] = "Última Llegada:"
                hora = datetime.fromisoformat(ultima['timestamp']).strftime('%H:%M')
                ws[f'B{row}'] = f"{hora} ({ultima.get('nombre', '')} {ultima.get('apellido', '')})"
                row += 1
        
        # Seguridad
        if self.alertas:
            row += 1
            ws[f'A{row}'] = "SEGURIDAD"
            ws[f'A{row}'].font = self.font_subtitulo
            row += 1
            
            criticas = len([a for a in self.alertas if a.get('nivel') == 'CRITICO'])
            moderadas = len(self.alertas) - criticas
            
            ws[f'A{row}'] = "Alertas de Seguridad:"
            ws[f'B{row}'] = len(self.alertas)
            ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color="FF0000")
            row += 1
            
            ws[f'A{row}'] = "  - Críticas:"
            ws[f'B{row}'] = criticas
            row += 1
            
            ws[f'A{row}'] = "  - Moderadas:"
            ws[f'B{row}'] = moderadas
            row += 1
        
        # Sorteos
        if self.sorteos:
            row += 1
            ws[f'A{row}'] = "SORTEOS"
            ws[f'A{row}'].font = self.font_subtitulo
            row += 1
            
            ws[f'A{row}'] = "Total Sorteos:"
            ws[f'B{row}'] = len(self.sorteos)
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def crear_listado_invitados(self):
        """Hoja 2: Listado de Invitados"""
        ws = self.wb.create_sheet("👥 Listado Invitados")
        
        # Headers
        headers = ["#", "Apellido", "Nombre", "Mesa", "Presente", "Hora Llegada", "QR Code"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Datos
        invitados_ordenados = sorted(self.invitados, key=lambda x: (x.get('apellido', ''), x.get('nombre', '')))
        
        for idx, inv in enumerate(invitados_ordenados, 2):
            # Buscar hora de llegada
            hora_llegada = "-"
            acred = next((a for a in self.acreditaciones 
                         if a.get('invitado_id') == inv['id'] and a.get('tipo') == 'ingreso'), None)
            if acred:
                hora_llegada = datetime.fromisoformat(acred['timestamp']).strftime('%H:%M:%S')
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=inv.get('apellido', ''))
            ws.cell(row=idx, column=3, value=inv.get('nombre', ''))
            ws.cell(row=idx, column=4, value=inv.get('mesa', ''))
            ws.cell(row=idx, column=5, value="✓ SÍ" if inv.get('presente') else "✗ NO")
            ws.cell(row=idx, column=6, value=hora_llegada)
            ws.cell(row=idx, column=7, value=inv.get('qr_code', ''))
            
            # Formato
            for col in range(1, 8):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.border
                cell.font = self.font_normal
                
                if col == 5:  # Columna Presente
                    if inv.get('presente'):
                        cell.fill = self.fill_verde
                    else:
                        cell.fill = self.fill_gris
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
    
    def crear_historial_acreditaciones(self):
        """Hoja 3: Historial de Acreditaciones"""
        ws = self.wb.create_sheet("⏰ Historial")
        
        # Headers
        headers = ["#", "Apellido", "Nombre", "Mesa", "Tipo", "Fecha", "Hora", "Kiosco"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Datos (ordenados por timestamp)
        acreditaciones_ordenadas = sorted(self.acreditaciones, 
                                         key=lambda x: x.get('timestamp', ''))
        
        for idx, acred in enumerate(acreditaciones_ordenadas, 2):
            dt = datetime.fromisoformat(acred['timestamp'])
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=acred.get('apellido', ''))
            ws.cell(row=idx, column=3, value=acred.get('nombre', ''))
            ws.cell(row=idx, column=4, value=acred.get('mesa', ''))
            ws.cell(row=idx, column=5, value=acred.get('tipo', '').upper())
            ws.cell(row=idx, column=6, value=dt.strftime('%d/%m/%Y'))
            ws.cell(row=idx, column=7, value=dt.strftime('%H:%M:%S'))
            ws.cell(row=idx, column=8, value=acred.get('kiosco_id', 1))
            
            # Formato
            for col in range(1, 9):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.border
                cell.font = self.font_normal
                
                # Color según tipo
                if col == 5:
                    if acred.get('tipo') == 'ingreso':
                        cell.fill = self.fill_verde
                    else:
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
    
    def crear_alertas_seguridad(self):
        """Hoja 4: Actividades Sospechosas"""
        ws = self.wb.create_sheet("🚨 Seguridad")
        
        # Headers
        headers = ["#", "Apellido", "Nombre", "Mesa", "Tipo Alerta", "Hora", "Nivel"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Datos
        if not self.alertas:
            ws.cell(row=2, column=1, value="Sin alertas de seguridad")
            ws.merge_cells('A2:G2')
            ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
            return
        
        for idx, alerta in enumerate(self.alertas, 2):
            dt = datetime.fromisoformat(alerta['timestamp'])
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=alerta.get('apellido', ''))
            ws.cell(row=idx, column=3, value=alerta.get('nombre', ''))
            ws.cell(row=idx, column=4, value=alerta.get('mesa', ''))
            ws.cell(row=idx, column=5, value=alerta.get('razon', ''))
            ws.cell(row=idx, column=6, value=dt.strftime('%H:%M:%S'))
            
            nivel = alerta.get('nivel', 'MEDIO')
            ws.cell(row=idx, column=7, value=f"🚨 {nivel}" if nivel == 'CRITICO' else f"⚠️ {nivel}")
            
            # Formato
            for col in range(1, 8):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.border
                cell.font = self.font_normal
                
                if nivel == 'CRITICO':
                    cell.fill = self.fill_rojo
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def crear_estadisticas_mesas(self):
        """Hoja 5: Estadísticas por Mesa"""
        ws = self.wb.create_sheet("📊 Por Mesa")
        
        # Agrupar por mesa
        mesas = {}
        for inv in self.invitados:
            mesa = inv.get('mesa', 'Sin mesa')
            if mesa not in mesas:
                mesas[mesa] = {'total': 0, 'presentes': 0, 'llegadas': []}
            
            mesas[mesa]['total'] += 1
            if inv.get('presente'):
                mesas[mesa]['presentes'] += 1
                
                # Buscar hora
                acred = next((a for a in self.acreditaciones 
                            if a.get('invitado_id') == inv['id'] and a.get('tipo') == 'ingreso'), None)
                if acred:
                    hora = datetime.fromisoformat(acred['timestamp'])
                    mesas[mesa]['llegadas'].append(hora)
        
        # Headers
        headers = ["Mesa", "Invitados", "Presentes", "Ausentes", "% Asistencia", "Hora Promedio"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Datos
        for idx, (mesa, datos) in enumerate(sorted(mesas.items()), 2):
            ausentes = datos['total'] - datos['presentes']
            pct = (datos['presentes'] / datos['total'] * 100) if datos['total'] > 0 else 0
            
            # Hora promedio
            hora_prom = "-"
            if datos['llegadas']:
                prom_seconds = sum(h.hour * 3600 + h.minute * 60 + h.second for h in datos['llegadas']) / len(datos['llegadas'])
                hora_prom = f"{int(prom_seconds // 3600):02d}:{int((prom_seconds % 3600) // 60):02d}"
            
            ws.cell(row=idx, column=1, value=mesa)
            ws.cell(row=idx, column=2, value=datos['total'])
            ws.cell(row=idx, column=3, value=datos['presentes'])
            ws.cell(row=idx, column=4, value=ausentes)
            ws.cell(row=idx, column=5, value=f"{pct:.0f}%")
            ws.cell(row=idx, column=6, value=hora_prom)
            
            # Formato
            for col in range(1, 7):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.border
                cell.font = self.font_normal
                cell.alignment = Alignment(horizontal='center')
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
    
    def crear_ganadores_sorteos(self):
        """Hoja 6: Ganadores de Sorteos"""
        ws = self.wb.create_sheet("🎰 Sorteos")
        
        # Headers
        headers = ["#", "Apellido", "Nombre", "Mesa", "Tipo Sorteo", "Hora"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Datos
        for idx, sorteo in enumerate(self.sorteos, 2):
            dt = datetime.fromisoformat(sorteo['fecha_sorteo'])
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=sorteo.get('apellido', ''))
            ws.cell(row=idx, column=3, value=sorteo.get('nombre', ''))
            ws.cell(row=idx, column=4, value=sorteo.get('mesa', ''))
            ws.cell(row=idx, column=5, value=sorteo.get('tipo', ''))
            ws.cell(row=idx, column=6, value=dt.strftime('%H:%M:%S'))
            
            # Formato
            for col in range(1, 7):
                cell = ws.cell(row=idx, column=col)
                cell.border = self.border
                cell.font = self.font_normal
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
